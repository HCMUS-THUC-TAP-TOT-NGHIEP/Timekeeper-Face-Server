import io
import os
import shutil
import threading
from datetime import date, datetime, time, timedelta
from threading import Thread

import numpy as np
from flask import current_app as app
from flask import url_for
from marshmallow import fields
from openpyxl.styles import (Alignment, Border, Color, Font, NamedStyle,
                             PatternFill, Side)
from pandas import DataFrame, ExcelWriter, option_context, set_option
from sqlalchemy import (ARRAY, BigInteger, Boolean, Column, Date, DateTime,
                        Integer, Numeric, SmallInteger, String, Time, and_, or_,
                        between, delete, func)
from werkzeug.datastructures import ImmutableMultiDict
from werkzeug.utils import secure_filename
from xlsxwriter import Workbook

from src import marshmallow
from src.db import db
from src.department.model import DepartmentModel
from src.employee.model import EmployeeModel, vEmployeeModel
from src.employee_checkin.AttendanceStatistic import AttendanceStatisticV2
from src.shift.model import ShiftModel, vShiftAssignmentDetail
from src.utils.extension import ProjectException
from src.utils.helpers import DeleteFile, GetDayOfWeek, daterange, subtractTime


class Timesheet(db.Model):
    __tablename__ = "Timesheet"
    __table_args__ = {'extend_existing': True}

    Id = Column(BigInteger(), primary_key=True)
    Name = Column(String(), nullable=False)
    DateFrom = Column(Date(), nullable=False)
    DateTo = Column(Date(), nullable=False)
    DepartmentList = Column(ARRAY(Integer()), default=[])
    LockedStatus = Column(Boolean(), default=False)
    CreatedBy = Column(Integer())
    CreatedAt = Column(DateTime(), default=datetime.now())
    ModifiedBy = Column(Integer())
    ModifiedAt = Column(DateTime(), default=datetime.now())

    def __init__(self) -> None:
        super().__init__()

    def __str__(self):
        return self.Name.replace('/', '-')

    def Lock(self):
        app.logger.info(f"Lock bảng công chi tiết mã {self.Id}")
        self.LockedStatus = True
        db.session.commit()

    def InsertTimesheetDetail(self) -> None:
        try:
            # region query: ghi lại chi tiết chấm công
            app.logger.info(f"Timesheet.InsertTimesheetDetail start. ")
            if self.LockedStatus:
                raise ProjectException(
                    f"Bảng chấm công mã {self.Id} - {self.Name} không thể chỉnh sửa vì đã khóa. ")
            DateFrom = self.DateFrom
            DateTo = self.DateTo
            DepartmentList = self.DepartmentList
            checkinRecordList = AttendanceStatisticV2.QueryMany(
                DateFrom=DateFrom, DateTo=DateTo)["items"]
            query = db.select(EmployeeModel)
            if len(DepartmentList) > 0:
                query = query.where(
                    EmployeeModel.DepartmentId.in_(DepartmentList))
            employeeList = db.session.execute(query).scalars()
            db.session.execute(db.delete(TimesheetDetail).where(
                TimesheetDetail.TimesheetId == self.Id))
            delta = timedelta(days=1)
            for employee in employeeList:
                currentDate = DateFrom
                endDate = DateTo
                while currentDate <= endDate:
                    timeSheetDetail = TimesheetDetail()
                    checkinRecords = list(filter(
                        lambda x: x.Id == employee.Id and x.Date == currentDate, checkinRecordList))
                    counts = checkinRecords.__len__()
                    timeSheetDetail.EmployeeId = employee.Id
                    timeSheetDetail.Date = currentDate
                    timeSheetDetail.TimesheetId = self.Id
                    if timeSheetDetail.IncludeAssignment():
                        timeSheetDetail.CheckinTime = checkinRecords[0].Time.time(
                        ) if counts > 0 else None
                        timeSheetDetail.CheckoutTime = checkinRecords[1].Time.time(
                        ) if counts > 1 else None
                        db.session.add(timeSheetDetail)
                    currentDate += delta
            # endregion
            return
        except ProjectException as pEx:
            app.logger.exception(
                f"Timesheet.InsertTimesheetDetail failed. Exception[{pEx}]")
            raise pEx
        except Exception as ex:
            app.logger.exception(
                f"Timesheet.InsertTimesheetDetail failed. Exception[{ex}]")
            raise Exception(
                f"Timesheet.InsertTimesheetDetail failed. Exception[{ex}]")
        finally:
            app.logger.exception(f"Timesheet.InsertTimesheetDetail finish. ")

    def QueryDetails(self, SearchString: str = "", DepartmentList: list() = [], EmployeeId=None,  Date=None):
        try:
            app.logger.info(f"Timesheet.QueryDetails start. ")
            query = db.select(vTimesheetDetail).where(
                vTimesheetDetail.TimesheetId == self.Id)
            if SearchString and SearchString.strip():
                sqlStr = "%"+SearchString.strip().upper()+"%"
                query = query.where(or_(func.cast(vTimesheetDetail.EmployeeId, String).like(
                    sqlStr), func.upper(vTimesheetDetail.EmployeeName).like(sqlStr)))
            if DepartmentList and len(DepartmentList):
                query = query.where(
                    vTimesheetDetail.Department.in_(DepartmentList))
            query = query.order_by(
                vTimesheetDetail.Department, vTimesheetDetail.EmployeeId)
            data = db.session.execute(query).scalars().all()
            return data
        except Exception as ex:
            app.logger.exception(
                f"Timesheet.QueryDetails failed. Exception[{ex}]")
            raise Exception(
                f"Timesheet.QueryDetails failed. Exception[{ex}]")
        finally:
            app.logger.exception(f"Timesheet.QueryDetails finish. ")

    def CreateTimesheetReport(self) -> str:
        try:
            app.logger.info(f"Timesheet.CreateTimesheetReport() start.")
            templatePath = os.path.join(
                os.pardir, app.static_folder, "templates", "Excel", f'SummaryReport.xlsx')
            path = os.path.join(os.pardir, app.static_folder, "export",
                                f'Bảng chấm công tổng hợp_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx')
            shutil.copyfile(templatePath, path)
            writer = ExcelWriter(path, engine="openpyxl",
                                 mode="a", if_sheet_exists="overlay")
            worksheet = writer.sheets["Main"]
            date_from = self.DateFrom.__format__("%d/%m/%Y")
            date_to = self.DateTo.__format__("%d/%m/%Y")
            worksheet["A3"] = f"Từ ngày {date_from} đến ngày {date_to}"

            # region write các date header

            daysColumns = [[date.day.__str__() for date in daterange(self.DateFrom, self.DateTo)], [
                GetDayOfWeek(date) for date in daterange(self.DateFrom, self.DateTo)]]
            df = DataFrame(data=daysColumns)
            df.to_excel(writer, sheet_name="Main", header=False,
                        startcol=14, startrow=4, index=False)
            colStyle = NamedStyle(name="colStyle")
            colStyle.fill = PatternFill("solid", fgColor=Color(index=44))
            colStyle.font = Font(name='Tahoma', size=8, bold=True)
            colStyle.border = Border(top=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"), left=Side(border_style="thin", color="000000")
                                     )
            colStyle.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
            for cells in worksheet.iter_rows(min_row=5, max_row=6, min_col=15):
                for cell in cells:
                    cell.style = colStyle
            # endregion

            # region write data

            start_row = 6
            start_col = 0
            stt = 1
            data = []
            query = db.select(vEmployeeModel)
            if self.DepartmentList:
                query = query.where(
                    vEmployeeModel.DepartmentId.in_(self.DepartmentList))
            #     departmentList = DepartmentList
            # if departmentList:
            query = query.order_by(
                vEmployeeModel.DepartmentId, vEmployeeModel.Id)
            employeeList = db.session.execute(query).scalars().all()
            if employeeList:
                for employee in employeeList:
                    row = [stt, employee.Id, employee.FullName(), employee.DepartmentName,
                           employee.Position, employee.JoinDate.strftime("%d/%m/%Y"), 0, 0, 0, 0, 0]
                    total_minute = 0
                    count_late_early = 0
                    count_no_timekeeping = 0
                    details = []
                    for date in daterange(self.DateFrom, self.DateTo):
                        record = db.session.execute(db.select(vTimesheetDetail).where(and_(
                            vTimesheetDetail.TimesheetId == self.Id, vTimesheetDetail.EmployeeId == employee.Id, vTimesheetDetail.Date == date))).scalars().first()
                        detailString = None
                        if not record:
                            continue
                        else:
                            detail = TimesheetDetailSchema().dump(record)
                            if not detail["CheckinTime"] and not detail["CheckoutTime"]:
                                count_no_timekeeping += 1
                            if detail["LateMinutes"] > 0:
                                count_late_early += 1
                                total_minute += round(detail["LateMinutes"], 2)
                            if detail["EarlyMinutes"] > 0:
                                count_late_early += 1
                                total_minute += round(
                                    detail["EarlyMinutes"], 2)
                            shiftName = detail["ShiftName"] if detail["ShiftName"] else ""
                            totalHour = round(detail["TotalHour"], 2).__str__(
                            ) + " h" if detail["TotalHour"] else ""
                            detailString = "\n".join([shiftName, totalHour])
                        details.append(detailString)
                    row.append(count_late_early)
                    row.append(total_minute)
                    row.append(count_no_timekeeping)
                    row.extend(details)
                    data.append(row)
                    stt += 1
            dataStyle = NamedStyle(name="dataStyle")
            dataStyle.font = Font(name='Tahoma', size=8)
            dataStyle.border = Border(top=Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"), bottom=Side(border_style="thin", color="000000"), left=Side(border_style="thin", color="000000")
                                      )
            dataStyle.alignment = Alignment(
                vertical="center", wrap_text=True, shrink_to_fit=True)
            df = DataFrame(data=data)
            df.to_excel(writer,  sheet_name="Main", startcol=start_col, startrow=start_row,
                        float_format="%.2f", index=False, header=False, na_rep="")
            for cells in worksheet.iter_rows(min_row=7, min_col=0):
                for cell in cells:
                    cell.style = dataStyle

            # endregion

            return path
        except ProjectException as pEx:
            app.logger.exception(
                f"Timesheet.CreateTimesheetReport() exception[{pEx}]")
            raise pEx
        except Exception as ex:
            app.logger.exception(
                f"Timesheet.CreateTimesheetReport() exception[{ex}]")
            raise ex
        finally:
            writer.close()
            app.logger.info(f"Timesheet.CreateTimesheetReport() finished.")

    @staticmethod
    def QueryMany(LockedStatus=None, Keyword=None, Page=None, PageSize=None):
        try:
            app.logger.info(f"Timesheet.QueryMany start. ")
            condition = ""
            if Keyword and Keyword.strip() != "":
                Keyword = Keyword.strip().lower()
                condition = func.lower(Timesheet.Name).like(f'%{Keyword}%')
            if LockedStatus:
                condition = and_(
                    condition, Timesheet.LockedStatus == LockedStatus)
            query = db.select(Timesheet)
            if condition:
                query = query.where(condition)
            if not Page:
                Page = 1
            if not PageSize:
                PageSize = 50
            data = db.paginate(query, page=Page, per_page=PageSize)
            return data
        except Exception as ex:
            app.logger.exception(
                f"Timesheet.QueryMany failed. Exception[{ex}]")
            raise Exception(
                f"Timesheet.QueryMany failed. Exception[{ex}]")
        finally:
            app.logger.exception(f"Timesheet.QueryMany finish. ")

    @staticmethod
    def DeleteById(id: int = None) -> bool:
        try:
            if id is None:
                raise ProjectException(
                    "Yêu cầu không hợp lệ, do chưa cung cấp mã báo cáo.")
            timesheet = Timesheet.query.filter_by(Id=id).first()
            if not timesheet:
                raise ProjectException(
                    f"Yêu cầu không thành công, do không tìm thấy bảng chấm công mã {id}.")
            db.session.execute(delete(TimesheetDetail).where(
                TimesheetDetail.TimesheetId == id))
            db.session.delete(timesheet)
            db.session.commit()
            return True
        except ProjectException as pEx:
            db.session.rollback()
            app.logger.exception(
                f"Timesheet.DeleteById thất bại. Có exception[{str(pEx)}]")
            raise pEx
        except Exception as ex:
            db.session.rollback()
            app.logger.exception(
                f"Timesheet.DeleteById thất bại. Có exception[{str(ex)}]")
            raise Exception(
                f"Timesheet.DeleteById thất bại. Có exception[{str(ex)}]")
        finally:
            app.logger.info(f"Timesheet.DeleteById kết thúc")


class TimesheetSchema(marshmallow.Schema):
    Id = fields.Integer()
    Name = fields.String()
    DateFrom = fields.Date()
    DateTo = fields.Date()
    DepartmentList = fields.List(fields.Integer())
    LockedStatus = fields.Boolean()


class TimesheetDetail(db.Model):
    __tablename__ = "TimesheetDetail"
    __table_args__ = {'extend_existing': True}

    Id = Column(BigInteger(), primary_key=True)
    TimesheetId = Column(BigInteger(), nullable=False)
    EmployeeId = Column(Integer(), nullable=False)
    Date = Column(Date(), nullable=False)
    CheckinTime = Column(Time())
    CheckoutTime = Column(Time())
    ShiftAssignmentId = Column(Integer())
    ShiftId = Column(Integer())
    StartTime = Column(Time())
    FinishTime = Column(Time())
    BreakAt = Column(Time())
    BreakEnd = Column(Time())
    LateMinutes = Column(Integer(), default=0)
    EarlyMinutes = Column(Integer(), default=0)

    def __init__(self) -> None:
        super().__init__()

    def IncludeAssignment(self) -> bool:
        try:
            app.logger.exception(f"TimesheetDetail.IncludeAssignment start. ")
            query = db.select(vShiftAssignmentDetail).where(and_(vShiftAssignmentDetail.EmployeeId == self.EmployeeId,
                                                                 between(
                                                                     self.Date, vShiftAssignmentDetail.StartDate, vShiftAssignmentDetail.EndDate)
                                                                 ))

            shiftAssignments = db.session.execute(query).scalars().all()

            for shiftAssignment in shiftAssignments:
                if shiftAssignment:
                    if self.Date.isoweekday() not in shiftAssignment.DaysInWeek:
                        continue
                    self.ShiftAssignmentId = shiftAssignment.Id
                    self.ShiftId = shiftAssignment.ShiftId
                    self.BreakAt = shiftAssignment.BreakAt
                    self.BreakEnd = shiftAssignment.BreakEnd
                    self.StartTime = shiftAssignment.StartTime
                    self.FinishTime = shiftAssignment.FinishTime
                    break
                # if self.CheckinTime:
                #     hourDiff = self.CheckinTime.hour - shiftAssignment.StartTime.hour
                #     minDiff = self.CheckinTime.minute - shiftAssignment.StartTime.minute
                #     lateMinutes = hourDiff*60 + minDiff
                #     self.LateMinutes = lateMinutes
                # if self.CheckoutTime:
                #     earlyMinutes = datetime.combine(date.today(
                #     ), shiftAssignment.FinishTime) - datetime.combine(date.today(), self.CheckoutTime)
                #     self.EarlyMinutes = earlyMinutes.total_seconds() / 60
            return True
        except Exception as ex:
            app.logger.exception(
                f"TimesheetDetail.IncludeAssignment failed. Exception[{ex}]")
            raise Exception(
                f"TimesheetDetail.IncludeAssignment failed. Exception[{ex}]")
        finally:
            app.logger.exception(f"TimesheetDetail.IncludeAssignment finish. ")

    def UpdateOne(self, data: dict = None):
        try:
            if data is None:
                return
            if "CheckinTime" in data and data["CheckinTime"] != self.CheckinTime:
                self.CheckinTime = data["CheckinTime"]
            if "CheckoutTime" in data and data["CheckoutTime"] != self.CheckoutTime:
                self.CheckoutTime = data["CheckoutTime"]
            db.session.commit()
            return
        except Exception as ex:
            db.session.rollback()
            app.logger.exception(
                f"TimesheetDetail.IncludeAssignment failed. Exception[{ex}]")
            raise Exception(
                f"TimesheetDetail.IncludeAssignment failed. Exception[{ex}]")
        finally:
            app.logger.exception(f"TimesheetDetail.IncludeAssignment finish. ")


class vTimesheetDetail(db.Model):
    __tablename__ = "vTimesheetDetail"
    __table_args__ = {'extend_existing': True}

    Id = Column(BigInteger(), primary_key=True)
    Date = Column(Date(), nullable=False)
    TimesheetId = Column(BigInteger(), nullable=False)
    EmployeeId = Column(Integer(), nullable=False)
    EmployeeName = Column(String())
    Department = Column(String())
    CheckinTime = Column(Time())
    CheckoutTime = Column(Time())
    ShiftAssignmentId = Column(Integer())
    ShiftId = Column(Integer())
    StartTime = Column(Time())
    FinishTime = Column(Time())
    BreakAt = Column(Time())
    BreakEnd = Column(Time())
    ShiftName = Column(String(), default="")
    LateMinutes = Column(Integer(), default=0)
    EarlyMinutes = Column(Integer(), default=0)
    DaysInWeek = Column(ARRAY(Integer()))
    WorkingHour = Column(Numeric(precision=10, scale=2))
    WorkingDay = Column(Numeric(precision=10, scale=2))

    def __init__(self) -> None:
        super().__init__()


class TimesheetDetailSchema(marshmallow.Schema):
    Id = fields.Integer()
    TimesheetId = fields.Integer()
    EmployeeId = fields.Integer()
    Date = fields.Date()
    CheckinTime = fields.Time()
    CheckoutTime = fields.Time()
    StartTime = fields.Time()
    FinishTime = fields.Time()
    BreakAt = fields.Time()
    BreakEnd = fields.Time()
    ShiftAssignmentId = fields.Integer()
    ShiftId = fields.Integer()
    EmployeeName = fields.String()
    Department = fields.String()
    ShiftName = fields.String()
    WorkingHour = fields.Integer()
    WorkingDay = fields.Integer()
    LateMinutes = fields.Method("GetLateMinutes")
    EarlyMinutes = fields.Method("GetEarlyMinutes")
    Status = fields.Method("GetStatus")
    TotalHour = fields.Method("GetTotalHours")
    RealWorkingHour = fields.Method("GetRealWorkingHours")

    def GetStatus(self, obj):
        try:
            if obj.CheckinTime < obj.StartTime:
                return "Vào sớm"
            else:
                return "Vào trễ"
        except:
            return None

    def GetLateMinutes(self, obj) -> int:
        try:
            if obj.StartTime and obj.CheckinTime:
                result = subtractTime(obj.StartTime, obj.CheckinTime)
                return result
            else:
                return 0
        except:
            return None

    def GetEarlyMinutes(self, obj) -> int:
        try:
            if obj.FinishTime and obj.CheckoutTime:
                result = subtractTime(obj.CheckoutTime, obj.FinishTime)
                return result
            else:
                return 0
        except:
            return None

    def GetTotalHours(self, obj) -> float:
        try:
            if not obj.BreakEnd or not obj.BreakAt:
                delta = subtractTime(obj.CheckinTime, obj.CheckoutTime)
                return delta / 60
            if (obj.CheckinTime > obj.BreakEnd):
                delta = subtractTime(obj.CheckinTime, obj.CheckoutTime)
            elif obj.CheckoutTime < obj.BreakEnd and obj.CheckinTime > obj.BreakAt:
                delta = subtractTime(obj.BreakEnd, obj.CheckoutTime)
            else:
                delta = subtractTime(
                    obj.CheckinTime, obj.CheckoutTime) - subtractTime(obj.BreakAt, obj.BreakEnd)
            return delta / 60
        except:
            return None

    def GetRealWorkingHours(self, obj) -> float:
        try:
            breakAt = obj.BreakAt
            breakEnd = obj.BreakEnd
            startTime = obj.StartTime
            finishTime = obj.FinishTime
            checkin = obj.CheckinTime
            checkout = obj.CheckoutTime
            if breakAt and breakEnd:
                hour1 = subtractTime(startTime, breakAt)
                if subtractTime(checkin, breakAt) < hour1:
                    hour1 = subtractTime(checkin, breakAt)
                if hour1 < 0:
                    hour1 = 0

                hour2 = subtractTime(breakEnd, checkout)
                if subtractTime(breakEnd, finishTime) < hour2:
                    hour2 = subtractTime(breakEnd, finishTime)
                if hour2 < 0:
                    hour2 = 0

                return (hour2 + hour1)/60
            if subtractTime(startTime, checkin) > 0:
                startTime = checkin
            if subtractTime(finishTime, checkout) < 0:
                finishTime = checkout
            return subtractTime(startTime, finishTime) / 60

        except Exception as ex:
            app.logger.exception(
                f"GetRealWorkingHours() failed. Exception[{ex}]")
            return 0


timesheetSchema = TimesheetSchema()
timesheetListSchema = TimesheetSchema(many=True)
timesheetDetailSchema = TimesheetDetailSchema()
timesheetDetailListSchema = TimesheetDetailSchema(many=True)
